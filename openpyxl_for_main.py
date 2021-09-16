import openpyxl


def createExcelFile(studentList): 
    wb = openpyxl.Workbook()
    ws = wb.active
    current = 1

    for student in studentList:
        ws['A{}'.format(current)] = student.firstName
        ws['B{}'.format(current)] = student.lastName
        ws['C{}'.format(current)] = student.grade
        ws['D{}'.format(current)] = student.gpa
        ws['E{}'.format(current)] = student.id
        current += 1


    wb.save('studentsExcelSheet.xlsx')
    
